#!/usr/bin/env python3
"""
bda_sync_pipeline.py

Pure-sync + threaded pipeline for:
- Veeva downloads (requests in ThreadPoolExecutor)
- BDA POST + GET polling (requests in ThreadPoolExecutor)
- Markdown extraction -> filter -> write text file
- Resume capability: Excel is source-of-truth; ARN column updated as jobs get posted

Edit CONFIGURATION section below before running.
"""

import os
import time
import json
import logging
import threading
import csv
from pathlib import Path
from typing import Dict, Any, Optional, Tuple, List
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
import pandas as pd
from openpyxl import load_workbook
from rapidfuzz import fuzz

# -----------------------
# CONFIGURATION (edit)
# -----------------------
EXCEL_PATH = "documents.xlsx"      # your excel
SHEET_NAME = None                  # default sheet or specify name
DOC_ID_COL = "documentid"          # exact column name for document id
ARN_COL = "arn"                    # column for arn (will be filled)
STATUS_COL = "status"              # status column (created if missing)

OUTPUT_DIR = Path("bda_outputs")   # where <docid>.txt files will be written
FAILED_CSV = "failed.csv"          # records failures

# Veeva
VEEVA_BASE_URL = os.environ.get("VEEVA_BASE_URL", "https://veeva.example.com")
VEEVA_BEARER_TOKEN = os.environ.get("VEEVA_BEARER_TOKEN", "YOUR_VEEVA_BEARER_TOKEN")
VEEVA_FETCH_TIMEOUT = 40           # seconds
VEEVA_WORKERS = 8                  # parallel Veeva threads (tune: 5-15 recommended)
VEEVA_RETRIES = 3

# BDA
BDA_POST_URL = os.environ.get("BDA_POST_URL", "https://bda.example.com/v1/recognize/bda")
BDA_GET_URL = os.environ.get("BDA_GET_URL", "https://bda.example.com/v1/recognize/bda")
BDA_API_KEY = os.environ.get("BDA_API_KEY", "YOUR_BDA_API_KEY")
BDA_WORKERS = 10                    # threads for BDA POST+poll
POLL_INTERVALS = [2, 3, 5, 8, 12, 20]  # progressive polling intervals
POLL_MAX_WAIT = 300                 # max seconds to wait before TIMEOUT

# Misc
MAX_TOTAL_THREADS = max(VEEVA_WORKERS, BDA_WORKERS)
EXCEL_WRITE_LOCK = threading.Lock()
LOG_LEVEL = logging.INFO

# -----------------------
# Logging
# -----------------------
logging.basicConfig(
    level=LOG_LEVEL,
    format="%(asctime)s %(levelname)s %(message)s"
)
logger = logging.getLogger("bda_sync_pipeline")

# -----------------------
# Utilities
# -----------------------
def ensure_output_dir():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def read_excel_to_df(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=SHEET_NAME, engine="openpyxl", dtype=str)
    # ensure required cols exist
    if DOC_ID_COL not in df.columns:
        raise RuntimeError(f"Excel missing required column '{DOC_ID_COL}'")
    if ARN_COL not in df.columns:
        df[ARN_COL] = ""
    if STATUS_COL not in df.columns:
        df[STATUS_COL] = ""
    return df

def save_cell_to_excel(path: str, row_index: int, col_name: str, value: str):
    """
    Write single cell to Excel (pandas index -> Excel row).
    Thread-safe via EXCEL_WRITE_LOCK.
    """
    with EXCEL_WRITE_LOCK:
        wb = load_workbook(path)
        ws = wb.active if SHEET_NAME is None else wb[SHEET_NAME]
        # map headers
        headers = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
        if col_name not in headers:
            # append col header
            last_col = ws.max_column + 1
            ws.cell(row=1, column=last_col, value=col_name)
            headers[col_name] = last_col
        excel_row = row_index + 2  # pandas idx 0 -> excel row2
        col_idx = headers[col_name]
        ws.cell(row=excel_row, column=col_idx, value=value)
        wb.save(path)

def append_failed_csv(doc_id: str, reason: str):
    write_header = not os.path.exists(FAILED_CSV)
    with threading.Lock():
        with open(FAILED_CSV, "a", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            if write_header:
                writer.writerow(["documentid", "reason", "timestamp"])
            writer.writerow([doc_id, reason, time.strftime("%Y-%m-%d %H:%M:%S")])

# -----------------------
# Veeva download (requests, retried)
# -----------------------
def fetch_file_from_veeva(doc_id: str) -> bytes:
    """
    Download binary file for doc_id. Retries internally.
    """
    url = f"{VEEVA_BASE_URL.rstrip('/')}/api/documents/{doc_id}/file"
    headers = {"Authorization": f"Bearer {VEEVA_BEARER_TOKEN}"}
    for attempt in range(1, VEEVA_RETRIES + 1):
        try:
            resp = requests.get(url, headers=headers, timeout=VEEVA_FETCH_TIMEOUT, stream=False)
            if resp.status_code == 200:
                return resp.content
            # handle auth errors explicitly
            if resp.status_code in (401, 403):
                msg = f"Auth error fetching {doc_id}: {resp.status_code}"
                logger.error(msg + " - Response: %s", resp.text[:200])
                raise RuntimeError(msg)
            if resp.status_code == 429:
                # rate limited, backoff
                wait = 5 * attempt
                logger.warning("Rate limited fetching %s (429). Backing off %ds", doc_id, wait)
                time.sleep(wait)
            else:
                logger.warning("Failed fetch %s status=%s", doc_id, resp.status_code)
        except Exception as e:
            logger.exception("Error fetching doc %s attempt %d: %s", doc_id, attempt, str(e))
        # exponential backoff
        time.sleep(2 * attempt)
    raise RuntimeError(f"Failed to fetch file for {doc_id} after retries")

# -----------------------
# BDA helpers (sync requests)
# -----------------------
def bda_post_file(file_bytes: bytes, model: str = "recognize-bda-latest") -> str:
    """
    POST file to BDA and return arn. Raises on error.
    """
    headers = {"Authorization": f"Bearer {BDA_API_KEY}"}
    files = {"file": ("file.pdf", file_bytes, "application/pdf")}
    data = {"model": model}
    resp = requests.post(BDA_POST_URL, headers=headers, files=files, data=data, timeout=180)
    resp.raise_for_status()
    j = resp.json()
    arn = j.get("invocation_arn") or j.get("arn") or j.get("job_id")
    if not arn:
        raise RuntimeError(f"No ARN returned in POST response: {j}")
    return arn

def bda_get_job(arn: str) -> Dict[str, Any]:
    headers = {"Authorization": f"Bearer {BDA_API_KEY}"}
    url = f"{BDA_GET_URL.rstrip('/')}/{arn}"
    resp = requests.get(url, headers=headers, timeout=60)
    resp.raise_for_status()
    return resp.json()

def poll_bda_until_done(arn: str, max_wait: int = POLL_MAX_WAIT) -> Optional[Dict[str, Any]]:
    waited = 0
    for interval in POLL_INTERVALS:
        try:
            j = bda_get_job(arn)
            status = (j.get("status") or j.get("job_status") or "").upper()
            if status == "COMPLETED":
                return j
            if status == "FAILED":
                return j
            logger.info("ARN %s status=%s; sleeping %ds", arn, status, interval)
        except Exception as e:
            logger.exception("Error polling ARN %s: %s", arn, e)
        time.sleep(interval)
        waited += interval
        if waited >= max_wait:
            break
    # final extended polling up to max_wait
    while waited < max_wait:
        try:
            j = bda_get_job(arn)
            status = (j.get("status") or j.get("job_status") or "").upper()
            if status == "COMPLETED":
                return j
            if status == "FAILED":
                return j
        except Exception as e:
            logger.exception("Final polling error ARN %s: %s", arn, e)
        time.sleep(10)
        waited += 10
    return None

# -----------------------
# Markdown extraction & filtering (plug your final filter here)
# -----------------------
def extract_markdown_from_bda_json(bda_json: Dict[str, Any]) -> str:
    md_list: List[str] = []
    def recur(o):
        if isinstance(o, dict):
            for k, v in o.items():
                if k == "markdown" and isinstance(v, str):
                    md_list.append(v)
                else:
                    recur(v)
        elif isinstance(o, list):
            for it in o:
                recur(it)
    recur(bda_json)
    return "\n\n".join(m.strip() for m in md_list if m and m.strip())

# Placeholder filter function: replace with your robust spaCy/scispacy filter
def filter_markdown(raw_markdown: str) -> str:
    if not raw_markdown:
        return ""
    # simple cleaning example - remove links and short lines
    txt = raw_markdown
    txt = re.sub(r"\[.*?\]\(.*?\)", "", txt)
    txt = re.sub(r"https?://\S+", "", txt)
    lines = []
    for ln in txt.splitlines():
        s = ln.strip()
        if not s:
            continue
        # drop typical marketing lines heuristically
        low = s.lower()
        if "click here" in low or "unsubscribe" in low or "prescribing information" in low:
            continue
        lines.append(s)
    return "\n".join(lines).strip()

# -----------------------
# Worker: per-document end-to-end
# -----------------------
def process_document(doc_id: str, row_idx: int, excel_path: str) -> None:
    """
    1) Download from Veeva
    2) POST to BDA -> write ARN to excel
    3) Poll GET until completion -> extract markdown
    4) Filter markdown -> write text file
    5) Update status column
    """
    try:
        logger.info("Starting doc %s", doc_id)
        # Fetch from Veeva
        file_bytes = fetch_file_from_veeva(doc_id)

        # POST to BDA with retries
        arn = None
        for attempt in range(1, 4):
            try:
                arn = bda_post_file(file_bytes)
                logger.info("Doc %s posted to BDA; ARN=%s", doc_id, arn)
                # write ARN to excel for resume
                save_cell_to_excel(excel_path, row_idx, ARN_COL, arn)
                save_cell_to_excel(excel_path, row_idx, STATUS_COL, "POSTED")
                break
            except Exception as e:
                logger.exception("BDA POST failed for %s attempt %d: %s", doc_id, attempt, e)
                time.sleep(2 * attempt)
        if not arn:
            append_failed_csv(doc_id, "BDA_POST_FAILED")
            save_cell_to_excel(excel_path, row_idx, STATUS_COL, "POST_FAILED")
            return

        # Poll until done
        result_json = poll_bda_until_done(arn)
        if result_json is None:
            append_failed_csv(doc_id, "BDA_TIMEOUT")
            save_cell_to_excel(excel_path, row_idx, STATUS_COL, "TIMEOUT")
            return

        status = (result_json.get("status") or result_json.get("job_status") or "").upper()
        if status == "FAILED":
            append_failed_csv(doc_id, "BDA_FAILED")
            save_cell_to_excel(excel_path, row_idx, STATUS_COL, "FAILED")
            # still try to extract whatever returned
        # Extract markdown and filter
        raw_md = extract_markdown_from_bda_json(result_json)
        filtered = filter_markdown(raw_md)

        # Save output
        ensure_output_dir()
        out_path = OUTPUT_DIR / f"{doc_id}.txt"
        with open(out_path, "w", encoding="utf-8") as fh:
            fh.write(filtered)

        save_cell_to_excel(excel_path, row_idx, STATUS_COL, "COMPLETED")
        logger.info("Completed doc %s; output=%s", doc_id, str(out_path))
    except Exception as e:
        logger.exception("Unhandled error processing %s: %s", doc_id, e)
        append_failed_csv(doc_id, f"UNHANDLED:{str(e)}")
        try:
            save_cell_to_excel(excel_path, row_idx, STATUS_COL, "ERROR")
        except Exception:
            pass

# -----------------------
# Orchestration
# -----------------------
def run_pipeline(excel_path: str):
    df = read_excel_to_df(excel_path)
    # Build list of tasks to process
    tasks: List[Tuple[str, int]] = []
    for idx, row in df.iterrows():
        doc_id = str(row[DOC_ID_COL]).strip()
        arn_val = str(row.get(ARN_COL, "") or "").strip()
        out_file = OUTPUT_DIR / f"{doc_id}.txt"
        # Resume logic: skip if completed already
        if arn_val and out_file.exists():
            logger.info("Skipping %s: ARN present and output exists", doc_id)
            continue
        # If ARN present but output missing, we will still attempt to poll using existing ARN:
        tasks.append((doc_id, idx))

    if not tasks:
        logger.info("No work to do. Exiting.")
        return

    logger.info("Processing %d documents with %d threads", len(tasks), MAX_TOTAL_THREADS)

    # ThreadPool: process documents in parallel with workers
    with ThreadPoolExecutor(max_workers=MAX_TOTAL_THREADS) as ex:
        futures = []
        for doc_id, idx in tasks:
            futures.append(ex.submit(process_document, doc_id, idx, excel_path))

        # progress as futures complete
        for fut in as_completed(futures):
            try:
                fut.result()
            except Exception as e:
                logger.exception("Worker raised exception: %s", e)

    logger.info("All tasks finished.")

# -----------------------
# Entrypoint
# -----------------------
if __name__ == "__main__":
    ensure_output_dir()
    if not os.path.exists(EXCEL_PATH):
        logger.error("Excel file not found: %s", EXCEL_PATH)
        raise SystemExit(1)
    run_pipeline(EXCEL_PATH)
